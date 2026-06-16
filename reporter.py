from datetime import date, timedelta
from pathlib import Path
from config import settings
from models import get_week_report, get_wrong_questions_top, get_all_users
from wecom_push import push_report_message


def get_week_number(d: date = None) -> tuple:
    if d is None:
        d = date.today()
    iso = d.isocalendar()
    return iso[0], iso[1]


def generate_week_report(year: int, week: int) -> str:
    report = get_week_report(year, week)
    users = report.get("users", [])
    total_users = len(get_all_users())
    participated = sum(1 for u in users if u["days_participated"] > 0)

    lines = [
        f"# 📊 星河十五五规划 · 第 {year} 年第 {week} 周成绩报告",
        f"**统计周期：** {report['start_date']} ~ {report['end_date']}",
        f"**本周考试天数：** {report['total_days']} 天",
        f"**应考人数：** {total_users} 人，**参与人数：** {participated} 人",
        f"**参与率：** {round(participated/total_users*100, 1) if total_users > 0 else 0}%",
        "",
    ]

    if users:
        lines.append("### 🏆 本周排名（Top 10）")
        lines.append("| 排名 | 姓名 | 部门 | 正确率 | 参与天数 |")
        lines.append("|------|------|------|--------|---------|")
        for i, u in enumerate(users[:10], 1):
            rate = round(u["correct_answers"] / u["total_answers"] * 100, 1) if u["total_answers"] > 0 else 0
            lines.append(f"| {i} | {u['name']} | {u['department']} | {rate}% | {u['days_participated']} |")

        lines.append("")
        lines.append("### 📋 个人完成情况")
        lines.append("| 姓名 | 部门 | 答题数 | 正确数 | 正确率 |")
        lines.append("|------|------|--------|--------|--------|")
        for u in users:
            rate = round(u["correct_answers"] / u["total_answers"] * 100, 1) if u["total_answers"] > 0 else 0
            lines.append(f"| {u['name']} | {u['department']} | {u['total_answers']} | {u['correct_answers']} | {rate}% |")

    wrong_questions = get_wrong_questions_top(10)
    if wrong_questions:
        lines.append("")
        lines.append("### ❌ 本周易错题 Top 10")
        for i, q in enumerate(wrong_questions, 1):
            opts = "  ".join([f"{k}. {v}" for k, v in q["options"].items()])
            lines.append(f"{i}. 第{q['day_seq']}天 第{q['seq']}题")
            lines.append(f"   {q['text']}")
            lines.append(f"   {opts}")
            lines.append(f"   ✅ 正确答案: **{q['correct_answer']}**  ")
            if q.get("explanation"):
                lines.append(f"   💡 解析: {q['explanation']}")
            lines.append(f"   ❌ 错误率: {q['wrong_rate']}% ({q['wrong_count']}/{q['total_count']}人答错)")

    return "\n".join(lines)


def generate_excel_report(year: int, week: int) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        print("请安装 openpyxl: pip install openpyxl")
        return Path()

    report = get_week_report(year, week)
    out_dir = Path(settings.report_output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    filename = out_dir / f"星河十五五_周报_{year}W{week:02d}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}年第{week}周"
    ws.merge_cells("A1:E1")
    title_cell = ws["A1"]
    title_cell.value = f"星河十五五规划 第{year}年第{week}周成绩报表"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal="center")

    headers = ["姓名", "部门", "答题数", "正确数", "正确率"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for i, u in enumerate(report["users"], 4):
        rate = round(u["correct_answers"] / u["total_answers"] * 100, 1) if u["total_answers"] > 0 else 0
        ws.cell(row=i, column=1, value=u["name"])
        ws.cell(row=i, column=2, value=u["department"])
        ws.cell(row=i, column=3, value=u["total_answers"])
        ws.cell(row=i, column=4, value=u["correct_answers"])
        ws.cell(row=i, column=5, value=f"{rate}%")

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    wb.save(str(filename))
    print(f"Excel 报表已生成: {filename}")
    return filename


def push_week_report():
    last_week = date.today() - timedelta(weeks=1)
    year, week = last_week.isocalendar()[0], last_week.isocalendar()[1]
    content = generate_week_report(year, week)
    success = push_report_message(content)
    if success:
        print(f"第 {year} 年第 {week} 周周报已推送")
    generate_excel_report(year, week)


def push_month_report():
    today = date.today()
    lines = [f"# 📊 {today.year}年{today.month}月成绩汇总", ""]
    for w in range(1, 5):
        dt = today - timedelta(weeks=w)
        yr, wk = dt.isocalendar()[0], dt.isocalendar()[1]
        report = get_week_report(yr, wk)
        users = report.get("users", [])
        lines.append(f"第{yr}年第{wk}周: {len(users)} 人参与")
    content = "\n".join(lines)
    push_report_message(content)


if __name__ == "__main__":
    import sys
    if "--week" in sys.argv:
        push_week_report()
    elif "--month" in sys.argv:
        push_month_report()
    else:
        year, week = get_week_number()
        print(generate_week_report(year, week))
